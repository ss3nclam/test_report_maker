from openpyxl import Workbook

from tb1_parser import AiSignal, ParsedTB1Collection, Signal, SignalsCollection

from .settings import SHEETS_CONFIG


class ReportMaker:

    def __init__(self, collection: ParsedTB1Collection) -> None:
        """
        docstring
        """
        if not isinstance(collection, ParsedTB1Collection):
            raise TypeError("некорректный тип аргумента")

        self.__logs_owner: str = self.__class__.__name__
        self.__collection: ParsedTB1Collection = collection.filter(
            key=lambda signal: signal.isused()
        )
        self.__wb = Workbook()

    # REFACT отрефакторить метод
    def __merge_identical_cells(self, sheet, columns_range: tuple = ("A", "B")):
        """
        docstring
        """
        ranges = []
        for column in columns_range:
            cells_list = tuple(cell.value for cell in sheet[column])
            unique_values_list = tuple({}.fromkeys(cells_list))[1:]
            unique_values_indexes = []

            for unique_value in unique_values_list:
                unique_values_indexes.append(cells_list.index(unique_value))

            unique_values_indexes.append(len(cells_list))

            for i, index_value in enumerate(unique_values_indexes):
                if index_value == unique_values_indexes[-1]:
                    break
                column_range: str = (
                    f"{column}{index_value + 1}:{column}{unique_values_indexes[i + 1]}"
                )
                ranges.append(column_range)

        for cell_range in ranges:
            sheet.merge_cells(cell_range)

    def __fill_Ai_sheet(self):
        """
        docstring
        """
        sheet = self.__wb["Ai"]
        for index, signal in enumerate(self.__collection["Ai"]):
            signal: AiSignal
            sheet.append((index + 1, signal.name, "знач."))
            sp_dict = {
                key: value
                for key, value in {
                    "НГ": signal.LL,
                    "НА": signal.LA,
                    "НП": signal.LW,
                    "ВП": signal.HW,
                    "ВА": signal.HA,
                    "ВГ": signal.HL,
                }.items()
                if value is not None
            }
            for sp_name, sp_value in sp_dict.items():
                sheet.append((index + 1, signal.name, sp_name, sp_value))
        self.__merge_identical_cells(sheet)

    def __fill_Prot_sheet(self):
        """
        docstring
        """
        sheet = self.__wb["Prot"]
        collection: ParsedTB1Collection | None = self.__collection.filter(
            key=lambda signal: signal.isprotected()
        )
        index = 1
        for key, value in collection.items():
            sheet.append([key])
            for signal in value:
                signal: Signal
                sheet.append((index, signal.name))
                index += 1

    def make_sheets(self):
        """
        docstring
        """
        for ws_name, ws_columns in SHEETS_CONFIG.items():
            self.__wb.create_sheet(ws_name) if ws_columns else None
            self.__wb[ws_name].append(ws_columns)
        del self.__wb["Sheet"]

        self.__fill_Ai_sheet()
        self.__fill_Prot_sheet()

    def write(self, filename: str):
        """
        docstring
        """
        self.__wb.save(filename)
